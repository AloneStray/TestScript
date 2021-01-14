'''
    封装openpyxl操作excel方法
    :create_time 20210104

'''
from openpyxl import load_workbook

class Openpyxl_Tool():

    #读操作
    def read_exceldata(self, filename=None, shtindex=None, start_row=None, start_col=None):
        '''
        读取已存在的excel指定行和列范围
        :param filename:
        :param shtindex:
        :param start_row:
        :param start_col:
        :return: 数据格式：[['',''],[]]
        '''
        wb = load_workbook(filename=filename)
        sht = wb.worksheets[shtindex]  # 根据索引选择工作表,从0开始

        data_all_g = sht.iter_rows(min_row=start_row, max=start_col)  # 按指定行读每一个值，行列索引从1开始,返回的是生成器

        # 加工返回的生成器格式：列表套列表，每行['','']  数据格式：[['',''],[]]
        data_list = []
        for row in data_all_g:
            list1 = []
            for cell in row:
                list1.append(cell.value)
            data_list.append(list1)

        return data_list


    #写操作
    def write_exceldata(self,datalists,filename=None,shtindex=None):
        '''
        已有sheet,按append(列表）形式按行追加，append从没有数据开始按行追加
        :return:
        '''
        wb = load_workbook(filename=filename)
        sht = wb.worksheets[shtindex]  # 根据索引选择工作表,从0开始
        #待补充优化：可以加判断，没有工作表就创建

        sht.append(datalists)
        wb.save(filename=filename)

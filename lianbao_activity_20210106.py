# 联报活动
import json
import requests
from openpyxl import load_workbook
import pymysql
import time

# 数据库操作
def conn_mysql():
    conn = pymysql.connect(host='XXX', user='xxx', passwd='xxx', db='xxx', port=xxx,
                           charset='utf8')
    return conn


def select_opreatemysql(SQL):
    conn = conn_mysql()
    cur = conn.cursor()  # 获取游标
    cur.execute(SQL)
    cur.close()  # 关闭游标
    conn.close()  # 关闭数据库
    return cur


# 读取excel 返回班级业务编号列表
def get_excel_data():
    wb = load_workbook(filename='联报活动.xlsx')
    sht = wb.worksheets[0]

    # 读取所有行的数据
    data_all = sht.iter_rows(min_row=7, min_col=1)
    data_list = list(data_all)

    classlists = []
    for row in data_list:
        tmp = []
        for cell in row:
            value = cell.value
            tmp.append(value)
        classlists.append(tmp)
    print(classlists)
    return classlists

# get_excel_data()

# 传入班级业务编号获取班级id列表
def get_clazznumbers():
    clazzbiznumbers = get_excel_data()
    clazznumbers_lists = []
    for row in clazzbiznumbers:
        tmp = []
        for i in row:

            SQL = "select number from gaotu.clazz  where biz_number = '%s'" % (i)
            cur = select_opreatemysql(SQL)
            value = cur.fetchone()  # 返回的是元组，需要转化成字符串
            result = ''
            for k in value:  # 元组转化成字符串
                result += str(k)
            tmp.append(result)

        clazznumbers_lists.append(tmp)
    return clazznumbers_lists  # 返回班级列表


# 构造请求数据subCoureInfo
def req_data():
    clazznumbers = get_clazznumbers()
    print(clazznumbers)
    lists = []
    for row_clazz in clazznumbers:
        tmp = []
        for v in row_clazz:
            subCoureInfo = {"clazzNumber": v, "activityPrice": 1}
            tmp.append(subCoureInfo)
        lists.append(tmp)
    return lists


# 构造联报活动请求
def active_method(reqdata):

    url = "https://xxx"
    header = {
        "content-type": "application/json;charset=UTF-8",
        "cookie": "xxx"
    }

    lb_name = "lbtest_" + str(int(time.time()*1000))
    data = {
        "title": lb_name,  # 修改为唯一字符串
        "capacity": 30,  # 写死
        "subCoureInfo": reqdata,

    # "subCoureInfo": [
    #     {
    #         "clazzNumber": "6299831369729024",
    #         "activityPrice": 1
    #     },
    #     {
    #         "clazzNumber": "6299822678017024",
    #         "activityPrice": 1
    #     }
    # ],

    "clazzTimeDesc": "2020年12月29日",  # 修改成当前日期
    "teacherNumber": "5999694498892800",  # 写死
    "assistantNumber": "5864668339335168",  # 写死
    "detailImgIds": [

    ]
    }
    print(json.dumps(data))
    res = requests.post(url=url, headers=header, json=data)
    return res.text,lb_name


#创建联报活动,联报号写入excel
def create_active():
    wb = load_workbook(filename='联报活动.xlsx')
    sht = wb.worksheets[1]

    data = req_data() #请求体的数据
    for i in data:
        results = active_method(i)
        print("返回本次请求结果：",results)
        lb_name = results[1]
        sql = "select number,name from gaotu.activity where name = '%s'"%(lb_name)
        cur = select_opreatemysql(sql)
        value = cur.fetchone()  # 返回的是元组，需要转化成字符串
        #print(value) #(6333975711318528, 'lbtest_1609729151385')

        if value != None:
            res_lists=[]
            for k in value:  # 元组转化成字符串
                result = ''
                result += str(k)
                res_lists.append(result)
            sht.append(res_lists)
    wb.save('联报活动.xlsx')

create_active()
'''
积累：
1.openpyxl读取excel
2.SQL拼接
'''